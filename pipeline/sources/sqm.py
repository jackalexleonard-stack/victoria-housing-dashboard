"""SQM Research rental vacancy rate — ``vic_vacancy`` (Melbourne + Regional Vic).

SQM's own site renders its vacancy data via JavaScript (the numbers aren't in the
served HTML), so instead we read the SQM-sourced vacancy series that the DFFH /
Homes Victoria Rental Report already publishes as "Figure 7: Rental vacancy rate"
— Metro Melbourne and Regional Victoria, monthly (13-term Henderson smoothed),
back to 1999. We reuse the DFFH workbook fetch, so this adds no new download.

Fig 7 source sheet layout: col 0 = Month (date), col 2 = Metro Melbourne %,
col 5 = Regional Victoria % (each region also has a raw-fraction column we skip).

Verified live 2026-07-16 (via the Sep-Qtr-2025 DFFH Rental Report workbook).
"""
from __future__ import annotations

import datetime as _dt
import io

import openpyxl
import pandas as pd

from pipeline import common
from pipeline.sources import dffh

_SHEET = "Fig 7 source"
# Percentage columns in the Fig 7 source sheet -> region.
_PCT_COLS = {2: "melbourne", 5: "regional_vic"}


def _num(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def parse_vacancy(raw: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    if _SHEET not in wb.sheetnames:
        raise ValueError(f"DFFH workbook has no '{_SHEET}' sheet")
    ws = wb[_SHEET]

    out = []
    for row in ws.iter_rows(values_only=True):
        d = row[0]
        if not isinstance(d, _dt.datetime):
            continue
        date = common.period_end(f"{d.year}-{d.month:02d}")
        for ci, region in _PCT_COLS.items():
            if ci < len(row) and _num(row[ci]):
                out.append((date, region, "vacancy_rate", float(row[ci]), "percent"))
    if not out:
        raise ValueError("no vacancy rows in DFFH Fig 7 source sheet")
    return pd.DataFrame(out, columns=common.TIDY_COLUMNS)


SERIES = [
    common.Series(
        id="vic_vacancy",
        source_name="Rental vacancy rate — SQM Research (via DFFH Rental Report)",
        source_url=dffh.INDEX,
        frequency="monthly",
        fetch=dffh.fetch_tables,
        parse=parse_vacancy,
    ),
]
