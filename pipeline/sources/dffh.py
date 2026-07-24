"""DFFH / Homes Victoria Rental Report (``vic_rents``).

Quarterly Victorian rental data, built from TWO workbooks linked off the same
report index (the quarter is baked into each URL slug, so we discover both
links fresh on each run rather than hardcoding them):

* "Tables from Rental Report - <Quarter> <Year>.xlsx" — ``fetch_tables`` /
  ``parse_tables``. Two genuine time series:
  * ``rent_growth_annual`` — Rent Index annual % change        (Fig 1 source, 2000Q2->)
  * ``affordable_share``   — affordable lettings % of new lets (Fig 8 source, 2020Q3->,
    published for Victoria/Metro/Regional — Victoria % is column 1, taken as-is,
    NOT derived from the Metro/Regional columns)
  (Table 1's current-quarter median-rent snapshot and Table 3's current-quarter
  dwelling-type snapshot are deliberately NOT read any more — both are
  superseded by full-history equivalents from the other workbook below.)
* "Quarterly median rents by Local Government Area.xlsx" — ``fetch_lga_medians`` /
  ``parse_lga_medians``. Every sheet shares the same METRO NON-METRO aggregate
  layout (Metro/Non-Metro/Victoria rows, one quarter-pair of columns each, back
  to Jun 1999, ~106 quarters). The Victoria row is a published statewide
  aggregate, not an average of the Metro/Non-Metro rows (these are medians —
  the mean of two medians is not a median) — taken as-is:
  * ``All Properties`` sheet -> ``median_rent`` — overall median rent, new lettings
  * the 6 dwelling-type sheets -> ``rent_<size>_<type>`` (one metric per sheet,
    see ``_LGA_DWELLING_SHEET`` for the sheet-name -> metric mapping; these are
    the exact same metric names Table 3 used to emit as a single-quarter
    snapshot, so nothing downstream that already reads them by name changes)
  Both come entirely from this workbook now. (Table 1's "Melbourne"/"Regional
  Victoria" median rent uses a slightly different statistical-region grouping
  than this workbook's LGA-based Metro/Non-Metro split — the two figures are
  close but not identical, so to avoid two competing sources for the same
  metric we no longer read Table 1 at all.)

``fetch_vic_rents``/``parse_vic_rents`` (used by the ``vic_rents`` Series below)
fetch and parse both workbooks and concatenate the tidy rows.

USER-AGENT NOTE: www.dffh.vic.gov.au tarpits/blocks non-browser User-Agents — the
plain pipeline UA times out, and even a browser UA with our identifier appended is
dropped; only a clean browser UA is served. So this one source overrides the
default UA out of necessity. We stay polite in every other respect: one run/day,
robots.txt permits these paths (/publications/ and the data files aren't
disallowed), and the usual timeouts + retries apply.

Discovery:  GET https://www.dffh.vic.gov.au/publications/rental-report
            -> href matching 'tables-rental-report-<quarter>-excel'
               and href matching 'quarterly-median-rents-local-government-area-<quarter>-excel'
            -> GET each URL (redirects to its .xlsx)
Verified live 2026-07-16 (September Quarter 2025 report; LGA-medians discovery
verified live 2026-07-21).
"""
from __future__ import annotations

import datetime as _dt
import io
import re

import openpyxl
import pandas as pd

from pipeline import common

BASE = "https://www.dffh.vic.gov.au"
INDEX = f"{BASE}/publications/rental-report"
# A clean browser UA — required; the site blocks anything else (see module docstring).
BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)
_TABLES_RE = re.compile(r'href="([^"]*tables-rental-report-[^"]*-excel)"', re.I)
_LGA_MEDIANS_RE = re.compile(
    r'href="([^"]*quarterly-median-rents-local-government-area-[^"]*-excel)"', re.I
)
_QUARTER_LABEL_FMT = "%b %Y"  # 'Jun 1999' as used by the LGA-medians header row


# --------------------------------------------------------------------------
# Fetch (browser UA; discover the current workbook link)
# --------------------------------------------------------------------------
def _browser_fetch(url: str) -> "object":
    return common.fetch(url, headers={"User-Agent": BROWSER_UA}, timeout=60)


def discover_tables_url(index_html: str) -> str:
    """Return the absolute URL of the current 'Tables from Rental Report' XLSX."""
    m = _TABLES_RE.search(index_html)
    if not m:
        raise ValueError("no 'tables-rental-report-*-excel' link on the report index")
    href = m.group(1)
    return href if href.startswith("http") else BASE + href


def discover_lga_medians_url(index_html: str) -> str:
    """Return the absolute URL of the current 'Quarterly median rents by LGA' XLSX."""
    m = _LGA_MEDIANS_RE.search(index_html)
    if not m:
        raise ValueError(
            "no 'quarterly-median-rents-local-government-area-*-excel' link "
            "on the report index"
        )
    href = m.group(1)
    return href if href.startswith("http") else BASE + href


def fetch_tables() -> bytes:
    html = _browser_fetch(INDEX).text
    return _browser_fetch(discover_tables_url(html)).content


def fetch_lga_medians() -> bytes:
    html = _browser_fetch(INDEX).text
    return _browser_fetch(discover_lga_medians_url(html)).content


def fetch_vic_rents() -> dict:
    """Fetch both workbooks behind ``vic_rents`` with a single INDEX page load."""
    html = _browser_fetch(INDEX).text
    return {
        "tables": _browser_fetch(discover_tables_url(html)).content,
        "lga_medians": _browser_fetch(discover_lga_medians_url(html)).content,
    }


# --------------------------------------------------------------------------
# Parse
# --------------------------------------------------------------------------
def _num(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _rows(ws) -> list:
    return list(ws.iter_rows(values_only=True))


def _timeseries(ws, colmap: dict[int, str], metric: str) -> list[tuple]:
    """Rows keyed by a datetime in col 0; ``colmap`` maps col index -> region.
    Values are fractions in the sheet and stored as percentages."""
    out = []
    for row in _rows(ws):
        d = row[0]
        if not isinstance(d, _dt.datetime):
            continue
        date = common.period_end(f"{d.year}-{d.month:02d}")
        for ci, region in colmap.items():
            if ci < len(row) and _num(row[ci]):
                out.append((date, region, metric, float(row[ci]) * 100.0, "percent"))
    return out


def parse_tables(raw: bytes) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

    rows: list[tuple] = []
    # Time series (metro = MRI / Metro col; regional = RRI / Regional col).
    # Table 1's overall median-rent snapshot and Table 3's dwelling-type
    # snapshot are deliberately NOT read here any more — ``parse_lga_medians``
    # below supplies full history for both from a different workbook, and
    # reading both would give those metrics two (slightly different) sources
    # for the same quarter.
    rows += _timeseries(wb["Fig 1 source"], {1: "melbourne", 2: "regional_vic"}, "rent_growth_annual")
    rows += _timeseries(
        wb["Fig 8 source"], {1: "vic", 2: "melbourne", 3: "regional_vic"}, "affordable_share"
    )

    return pd.DataFrame(rows, columns=common.TIDY_COLUMNS)


# --------------------------------------------------------------------------
# Parse: "Quarterly median rents by Local Government Area" workbook
# --------------------------------------------------------------------------
_LGA_SHEET = "All Properties"
_LGA_REGION = {"metro": "melbourne", "non-metro": "regional_vic", "victoria": "vic"}
# Dwelling-type sheet name (exact, as openpyxl reports it — casing is
# inconsistent in the source workbook, e.g. '1br flat' vs '2br Flat') -> the
# same tidy metric names Table 3 used to emit as a single-quarter snapshot.
_LGA_DWELLING_SHEET = {
    "1br flat": "rent_1br_flat",
    "2br Flat": "rent_2br_flat",
    "3br Flat": "rent_3br_flat",
    "2br House": "rent_2br_house",
    "3br House": "rent_3br_house",
    "4br House": "rent_4br_house",
}


def _quarter_label_to_iso(label) -> str:
    dt = _dt.datetime.strptime(str(label).strip(), _QUARTER_LABEL_FMT)
    return common.period_end(f"{dt.year}-{dt.month:02d}")


def _lga_quarter_columns(header_row: tuple) -> dict[int, str]:
    """Map each quarter's MEDIAN column index -> ISO period-end.

    Row 1 of the sheet gives each quarter a label spanning a (Count, Median)
    column pair, starting at column 2 (e.g. cols 2/3 = 'Jun 1999' Count/Median,
    cols 4/5 = 'Sep 1999' Count/Median, ...). We only need the Median column.
    """
    cols: dict[int, str] = {}
    for ci in range(2, len(header_row) - 1, 2):
        label = header_row[ci]
        if label is not None:
            cols[ci + 1] = _quarter_label_to_iso(label)
    return cols


def _lga_metro_trend(ws, metric: str) -> list[tuple]:
    """Label-scan one sheet of the LGA-medians workbook for the METRO
    NON-METRO section's Metro / Non-Metro / Victoria aggregate rows (NOT
    fixed row numbers — the sheet grows a suburb/LGA row occasionally), then
    walk every quarter-pair column to emit ``metric``'s full history for
    melbourne / regional_vic / vic. Shared by the ``All Properties`` sheet (-> ``median_rent``)
    and each of the 6 dwelling-type sheets (-> ``rent_<size>_<type>``) — same
    layout in every one of the 7 sheets.
    """
    rows = _rows(ws)
    quarter_cols = _lga_quarter_columns(rows[1])

    out: list[tuple] = []
    in_section = False
    for row in rows:
        label0 = str(row[0]).strip().lower() if row[0] is not None else ""
        if label0 == "metro non-metro":
            in_section = True
        if not in_section:
            continue
        label1 = str(row[1]).strip().lower() if row[1] is not None else ""
        region = _LGA_REGION.get(label1)
        if not region:
            continue
        for ci, date in quarter_cols.items():
            if ci < len(row) and _num(row[ci]):
                out.append((date, region, metric, float(row[ci]), "AUD/week"))
    return out


def parse_lga_medians(raw: bytes) -> pd.DataFrame:
    """'Quarterly median rents by Local Government Area' workbook: label-scan
    the ``All Properties`` sheet (-> ``median_rent``) and each of the 6
    dwelling-type sheets (-> ``rent_<size>_<type>``, see
    ``_LGA_DWELLING_SHEET``) for their Metro/Non-Metro aggregate rows, giving
    every one of these metrics the same ~106-quarter history back to Jun 1999.
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

    out: list[tuple] = list(_lga_metro_trend(wb[_LGA_SHEET], "median_rent"))
    for sheet_name, metric in _LGA_DWELLING_SHEET.items():
        out += _lga_metro_trend(wb[sheet_name], metric)

    if not out:
        raise ValueError(
            f"no Metro/Non-Metro median-rent rows found in '{_LGA_SHEET}' sheet"
        )
    return pd.DataFrame(out, columns=common.TIDY_COLUMNS)


def parse_vic_rents(raw: dict) -> pd.DataFrame:
    """Combine both workbooks' tidy rows for the ``vic_rents`` series."""
    df_tables = parse_tables(raw["tables"])
    df_lga = parse_lga_medians(raw["lga_medians"])
    return pd.concat([df_tables, df_lga], ignore_index=True)[common.TIDY_COLUMNS]


SERIES = [
    common.Series(
        id="vic_rents",
        source_name="DFFH / Homes Victoria Rental Report",
        source_url=INDEX,
        frequency="quarterly",
        fetch=fetch_vic_rents,
        parse=parse_vic_rents,
    ),
]
